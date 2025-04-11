import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

file = "../local-db/000000/questions.json"

def export_to_excel(data, output_file="../local-db/000000/questions.xlsx"):
    # 准备数据
    rows = []
    for item in data:
        chunk_id = item["chunkId"]
        for question in item["questions"]:
            rows.append({
                "chunkId": chunk_id,
                "question": question
            })
    
    # 创建DataFrame并导出到Excel
    df = pd.DataFrame(rows)
    df.to_excel(output_file, index=False)
    
    # 处理合并单元格
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 设置列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 80
    
    # 合并相同chunkId的单元格
    current_chunk = None
    start_row = 2  # 从第2行开始(第1行是标题)
    merge_ranges = []
    
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
        if row[0] != current_chunk:
            if current_chunk is not None and start_row < i-1:
                merge_ranges.append(f"A{start_row}:A{i-1}")
            current_chunk = row[0]
            start_row = i
    
    # 处理最后一组
    if current_chunk is not None and start_row < ws.max_row:
        merge_ranges.append(f"A{start_row}:A{ws.max_row}")
    
    # 执行合并
    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)
    
    # 设置垂直居中
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center')
    
    wb.save(output_file)
    print(f"Excel文件已生成: {output_file}")

with open(file) as f:
    chunks = json.load(f)
    export_to_excel(chunks)